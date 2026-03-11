import openpyxl
from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils.text import slugify

from docentes.models import Docente
from grupos.models import Grupo, DistribucionVacantes
from horarios.models import Horario
from asignaturas.models import Asignatura, Equivalencia
from aulas.models import Aula
import re

class ExcelImportError(Exception):
    pass

def importar_programacion(archivo_excel, user, periodo, escuela):

    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        ws = wb.active
    except Exception as e:
        raise ExcelImportError(f"El archivo no es un Excel válido. {str(e)}")

    header_row_idx = None
    col_map = {}
    vacantes_cols = {}

    # Encontramos la cabecere y mapeamos columnas (ASIGNATURA, GRUPO, DIA, H.INI, H.FIN, DOCENTE, TIPO, AULA)
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_str = [str(c).upper().strip() if c else '' for c in row]
        if "ASIGNATURA" in row_str and "GRUPO" in row_str:
            header_row_idx = row_idx

            # Detectar columnas base
            for col_idx, cell_value in enumerate(row_str):
                if cell_value:
                    col_map[cell_value] = col_idx

            # Detectar columnas de vacantes (después de AULA o DOCENTE)
            inicio_vacantes = col_map.get('AULA', col_map.get('DOCENTE', -1)) + 1
            if inicio_vacantes > 0:
                for idx in range(inicio_vacantes, len(row_str)):
                    sigla_escuela = row_str[idx]
                    if sigla_escuela:
                        vacantes_cols[idx] = sigla_escuela
            break

    if not header_row_idx:
        raise ExcelImportError("No se encontró la fila de cabecera con 'ASIGNATURA' y 'GRUPO'.")

    # Validar columnas
    required_cols = ['ASIGNATURA', 'GRUPO', 'DIA', 'H.INI', 'H.FIN', 'DOCENTE', 'T']
    for col in required_cols:
        if col not in col_map:
            raise ExcelImportError(f"Falta la columna obligatoria: {col}")

    if not escuela:
        raise ExcelImportError("Es necesario especificar una escuela para la importación.")

    # IDS de asignaturas que el usuario puede gestionar
    qs_asignaturas = _queryset_asignaturas_para_usuario(user, escuela)

    if not qs_asignaturas.exists():
        raise ExcelImportError(f"No se encontraron asignaturas para la escuela {escuela}.")

    errores = []
    registros_creados = 0

    with transaction.atomic(): # Si falla algo grave, no guarda nada
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
            try:
                # Asignatura
                idx_asig = col_map['ASIGNATURA']
                raw_asignatura = row[idx_asig]
                nombre_asignatura = str(raw_asignatura).strip() if raw_asignatura else None

                if not nombre_asignatura:
                    continue

                asignatura = qs_asignaturas.filter(nombre__iexact=nombre_asignatura).first()

                if not asignatura:
                    errores.append(f"Fila {row_idx}: Asignatura '{nombre_asignatura}' no encontrada en {escuela}.")
                    continue

                # Grupo
                idx_grupo = col_map['GRUPO']
                raw_grupo = row[idx_grupo]
                if isinstance(raw_grupo, float):
                    numero_grupo = str(int(raw_grupo))
                else:
                    numero_grupo = str(raw_grupo).replace('G', '').strip() if raw_grupo else None

                if not numero_grupo:
                    continue

                grupo, _ = Grupo.objects.get_or_create(
                    asignatura=asignatura,
                    numero=numero_grupo,
                    periodo=periodo,
                )

                vacantes_por_asignatura = {}

                for col_idx, sigla_escuela in vacantes_cols.items():
                    val_vacantes = row[col_idx]
                    if val_vacantes and str(val_vacantes).isdigit() and int(val_vacantes) > 0:
                        cantidad = int(val_vacantes)
                        asignatura_vacante = _resolver_asignatura_vacante(asignatura, sigla_escuela)
                        if asignatura_vacante:
                            if asignatura_vacante in vacantes_por_asignatura:
                                vacantes_por_asignatura[asignatura_vacante] += cantidad
                            else:
                                vacantes_por_asignatura[asignatura_vacante] = cantidad
                        else:
                            errores.append(f"Fila {row_idx}: No se pudo resolver equivalente para '{sigla_escuela}'.")

                for asig_obj, total_vacantes in vacantes_por_asignatura.items():
                    distribucion, created = DistribucionVacantes.objects.get_or_create(
                        grupo=grupo,
                        asignatura=asig_obj,
                        defaults={'cantidad': total_vacantes}
                    )
                    # Si ya existía y el total es diferente, lo actualizamos
                    if not created and distribucion.cantidad != total_vacantes:
                        distribucion.cantidad = total_vacantes
                        distribucion.save()

                # Docente
                idx_docente = col_map['DOCENTE']
                raw_docente = row[idx_docente]
                docente = _get_or_create_docente(raw_docente, user)

                # Horario
                idx_dia = col_map['DIA']
                idx_hini = col_map['H.INI']
                idx_hfin = col_map['H.FIN']
                idx_tipo = col_map['T']

                dia_raw = row[idx_dia]
                h_ini_raw = row[idx_hini]
                h_fin_raw = row[idx_hfin]
                tipo_raw = str(row[idx_tipo]).strip() if row[idx_tipo] else 'T'
                # Normalización
                dia_norm = _normalizar_dia(dia_raw)

                # Aula
                aula_raw = str(row[col_map['AULA']]).strip().upper() if row[col_map['AULA']] else ""
                aula_obj = None

                if aula_raw:
                    if '-' in aula_raw:
                        a_nom, p_raw = aula_raw.split('-', 1)
                    else:
                        a_nom, p_raw = aula_raw, 'NP'

                    p_code = p_raw.strip()
                    if p_code not in ['AP', 'NP']: p_code = 'NP'

                    aula_obj, _ = Aula.objects.get_or_create(
                        nombre=a_nom.strip(),
                        pabellon=p_code,
                        defaults={
                            'facultad': escuela.facultad,
                            'vacantes': 40,
                            'tipo_sesion': 'T' if str(row[col_map['T']]).upper().startswith('T') else 'P'
                        }
                    )

                if dia_norm and h_ini_raw and h_fin_raw:
                    horario, created = Horario.objects.get_or_create(
                        grupo=grupo, dia=dia_norm, hora_inicio=h_ini_raw,
                        defaults={
                            'hora_fin': h_fin_raw, 'tipo': tipo_raw[0].upper(),
                            'docente': docente, 'aula': aula_obj
                        }
                    )

                    if not created:
                        horario.hora_fin = h_fin_raw
                        horario.docente = docente
                        horario.aula = aula_obj
                        horario.save()

                    registros_creados += 1

                else:
                    errores.append(f"Fila {row_idx}: Datos de horario incompletos.")

            except Exception as e:
                errores.append(f"Fila {row_idx}: Error procesando datos - {str(e)}")

    return {"creados": registros_creados, "errores": errores}

def _queryset_asignaturas_para_usuario(user, escuela):

    qs = Asignatura.objects.filter(plan__escuela=escuela)

    rol_name = getattr(user.rol, 'name', None) if hasattr(user, 'rol') and user.rol else None

    if rol_name in ['Coordinador de Estudios Generales', 'Jefe de Estudios Generales']:
        qs = qs.filter(ciclo__in=[1, 2])

    return qs

def _resolver_asignatura_vacante(asignatura_base, sigla_escuela):

    MAPEO_SIGLAS_ESCUELA = {
        'SI': 'Ingeniería de Sistemas',
        'SW': 'Ingeniería de Software',
        'CC': 'Ciencias de la Computación',
        #'IA': 'INTELIGENCÍA ARTIFICIAL'
    }

    sigla_norm = str(sigla_escuela).strip().upper()
    nombre_escuela_buscada = MAPEO_SIGLAS_ESCUELA.get(sigla_norm)

    if not nombre_escuela_buscada:
        return None

    if hasattr(asignatura_base, 'plan') and asignatura_base.plan.escuela.nombre == nombre_escuela_buscada:
        return asignatura_base

    # Buscamos la tabla de equivalencias para la asignatura base
    equivalencia = Equivalencia.objects.filter(asignaturas=asignatura_base).first()

    if equivalencia:
        # Iteramos sobre los cursos equivalentes
        for asig_equivalente in equivalencia.asignaturas.all():
            if hasattr(asig_equivalente, 'plan') and asig_equivalente.plan.escuela.nombre == nombre_escuela_buscada:
                return asig_equivalente
    return None

def _get_or_create_docente(nombre_completo_raw, user):
    if not nombre_completo_raw:
        return None

    raw = str(nombre_completo_raw).strip().upper()

    paterno, materno, nombres = "", "", "SIN NOMBRE"

    if ',' in raw:
        # Formato: APpaterno APmaterno, Nombres
        partes = raw.split(',')
        apellidos = partes[0].strip().split()
        if len(partes) > 1:
            nombres = partes[1].strip()

        if apellidos:
            paterno = apellidos[0]
            materno = " ".join(apellidos[1:]) if len(apellidos) > 1 else ''

    else:
        # Formato: APpaterno APmaterno Nombres
        partes = raw.split()
        if len(partes) >= 2:
            nombres = partes[-1]
            apellidos_list = partes[:-1]
            paterno = apellidos_list[0]
            materno = " ".join(apellidos_list[1:]) if len(apellidos_list) > 1 else ''
        elif len(partes) == 1:
            paterno = partes[0]

    docente = Docente.objects.filter(
        apellido_paterno__iexact=paterno,
        nombres__iexact=nombres
    ).first()

    if not docente:
        facultad = getattr(user, 'facultad', None)
        if not facultad and hasattr(user, 'escuela') and user.escuela:
            facultad = user.escuela.facultad

        docente = Docente.objects.create(
            apellido_paterno=paterno,
            apellido_materno=materno,
            nombres=nombres,
            tipo='C', # Contratado
            facultad=facultad
        )

    return docente

def _normalizar_dia(dia_excel):
    dias = {
        'LUNES': '1', 'MARTES': '2', 'MIERCOLES': '3', 'JUEVES': '4',
        'VIERNES': '5', 'SABADO': '6', 'DOMINGO': '7',
        'LU': '1', 'MA': '2', 'MI': '3', 'JU': '4', 'VI': '5', 'SA': '6'
    }

    if not dia_excel:
        return None

    dia_str = str(dia_excel).upper().replace('É', 'E').strip()

    match = re.search(r'[A-Z]+', dia_str)
    if match:
        dia_str = match.group(0)

    return dias.get(dia_str, None)
