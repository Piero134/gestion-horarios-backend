import io
import hashlib
from collections import defaultdict
from itertools import groupby
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils.text import slugify
from random import randint

# Estilo del excel
class ExcelEstilos:
    # Colores
    COLOR_CABECERA_AZUL = "DCE6F1"   # Azul claro
    COLOR_CABECERA_TABLA = "F2F2F2"  # Gris claro
    COLOR_TITULO_CICLO = "BDD7EE"    # Azul medio

    # Bordes
    BORDER_THIN = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Fuentes
    FONT_TITLE_MAIN = Font(name='Calibri', size=11, bold=True)
    FONT_TITLE_SUB = Font(name='Calibri', size=11, bold=False)
    FONT_CICLO = Font(name='Calibri', size=10, bold=True)
    FONT_HEADER = Font(name='Calibri', size=9, bold=True)
    FONT_NORMAL = Font(name='Calibri', size=9)
    FONT_GRID = Font(name='Calibri', size=8, bold=True)

    # Alineación
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)


def generar_color_hex(nombre):
    nombre_normalizado = nombre.strip().upper()

    hash_obj = hashlib.md5(nombre_normalizado.encode())
    hash_hex = hash_obj.hexdigest()

    # Tomar partes del hash para RGB
    r = int(hash_hex[0:2], 16)
    g = int(hash_hex[2:4], 16)
    b = int(hash_hex[4:6], 16)

    # Convertir a tonos suaves (pastel)
    r = (r + 255) // 2
    g = (g + 255) // 2
    b = (b + 255) // 2

    return f"{r:02X}{g:02X}{b:02X}"

def _agregar_cabecera(ws, facultad_nombre, dependencia_nombre, nombre_periodo, ancho_total=12):
    partes = dependencia_nombre.split(" - ")

    linea_dependencia = partes[0]

    linea_escuela = partes[1] if len(partes) > 1 else ""

    lines = [
        "UNIVERSIDAD NACIONAL MAYOR DE SAN MARCOS",
        facultad_nombre,
        linea_dependencia,
        linea_escuela,
        f"HORARIO DE ESTUDIOS SEMESTRE ACADEMICO {nombre_periodo}"
    ]

    for i, line in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1)
        cell.value = line

        if i == 5:
            cell.font = ExcelEstilos.FONT_TITLE_MAIN
        else:
            cell.font = ExcelEstilos.FONT_TITLE_SUB

        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ancho_total)
        cell.alignment = ExcelEstilos.ALIGN_CENTER


# Hoja 1: Lista de grupos con sus horarios y docentes
def _generar_hoja_lista_cursos(wb, grupos, facultad_texto, dependencia_texto, nombre_periodo):
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Profesores")
    else:
        ws.title = "Profesores"

    current_row = 7
    ancho_tabla = 7

    _agregar_cabecera(ws, facultad_texto, dependencia_texto, nombre_periodo, ancho_total=ancho_tabla)

    grupos_por_ciclo = defaultdict(list) # Diccionario ciclo -> lista de grupos
    for g in grupos:
        grupos_por_ciclo[g.asignatura.ciclo].append(g)

    for ciclo in sorted(grupos_por_ciclo.keys()):
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ancho_tabla)
        cell = ws.cell(row=current_row, column=1, value=f"CICLO: {ciclo}")
        cell.fill = PatternFill(start_color=ExcelEstilos.COLOR_TITULO_CICLO, fill_type="solid")
        cell.font = ExcelEstilos.FONT_CICLO
        cell.alignment = ExcelEstilos.ALIGN_LEFT
        cell.border = ExcelEstilos.BORDER_THIN
        current_row += 2

        grupos_ciclo = grupos_por_ciclo[ciclo]
        grupos_ciclo.sort(key=lambda x: x.asignatura.nombre)

        for nombre_asig, grupos_iter in groupby(grupos_ciclo, key=lambda x: x.asignatura.nombre):
            grupos_asignatura = list(grupos_iter)
            grupo_base = grupos_asignatura[0]
            asig = grupo_base.asignatura

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ancho_tabla)
            cell = ws.cell(row=current_row, column=1, value=f"({asig.codigo}) {asig.nombre}")

            # Estilos
            cell.font = ExcelEstilos.FONT_TITLE_SUB
            cell.fill = PatternFill(start_color=ExcelEstilos.COLOR_CABECERA_AZUL, fill_type="solid")
            cell.alignment = ExcelEstilos.ALIGN_CENTER # Centrado
            cell.border = ExcelEstilos.BORDER_THIN

            # Aplicar borde a las celdas ocultas por el merge para asegurar el cuadro completo
            for i in range(2, ancho_tabla + 1):
                ws.cell(row=current_row, column=i).border = ExcelEstilos.BORDER_THIN

            current_row += 1

            # Cabeceras de la tabla de grupos
            #headers = ["GRUPO", "T", "DIA", "HORARIO", "DOCENTE", "AULA"]

            # Col 1: GRUPO
            c = ws.cell(row=current_row, column=1, value="GRUPO")
            # Col 2: T
            c = ws.cell(row=current_row, column=2, value="T")
            # Col 3: DIA
            c = ws.cell(row=current_row, column=3, value="DÍA")
            # Col 4: HORARIO
            c = ws.cell(row=current_row, column=4, value="HORARIO")
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=5).border = ExcelEstilos.BORDER_THIN # Borde celda oculta
            # Col 6: DOCENTE
            c = ws.cell(row=current_row, column=6, value="DOCENTE")
            # Col 7: AULA
            c = ws.cell(row=current_row, column=7, value="AULA")

            # Estilos de la fila de cabecera
            for i in range(1, ancho_tabla + 1):
                cell = ws.cell(row=current_row, column=i)
                cell.font = ExcelEstilos.FONT_HEADER
                cell.fill = PatternFill(start_color=ExcelEstilos.COLOR_CABECERA_TABLA, fill_type="solid")
                cell.border = ExcelEstilos.BORDER_THIN
                cell.alignment = ExcelEstilos.ALIGN_CENTER

            current_row += 1

            # Datos de los grupos
            for grp in grupos_asignatura:
                horarios = grp.horarios.all()
                fila_inicio_grupo = current_row
                iter_horarios = horarios if horarios else [None]

                for h in iter_horarios:
                    # Col 1: Grupo
                    ws.cell(row=current_row, column=1, value=f"G-{grp.numero}")

                    if h:
                        # Col 2: Tipo
                        ws.cell(row=current_row, column=2, value=h.tipo)
                        # Col 3: Dia
                        ws.cell(row=current_row, column=3, value=h.get_dia_display())
                        # Col 4: Inicio
                        ws.cell(row=current_row, column=4, value=h.hora_inicio.strftime('%H:%M'))
                        # Col 5: Fin
                        ws.cell(row=current_row, column=5, value=h.hora_fin.strftime('%H:%M'))
                        # Col 6: Docente
                        docente_nombre = str(h.docente) if h.docente else "POR ASIGNAR"
                        ws.cell(row=current_row, column=6, value=docente_nombre)
                        # Col 7: Aula
                        aula_nombre = h.aula.nombre if h.aula else "-"
                        ws.cell(row=current_row, column=7, value=aula_nombre)
                    else:
                        # Rellenar con guiones si no hay horario
                        for c_idx in range(2, 6): ws.cell(row=current_row, column=c_idx, value="-")
                        ws.cell(row=current_row, column=7, value="-")

                    # Estilos de fila
                    for c in range(1, ancho_tabla + 1):
                        cell = ws.cell(row=current_row, column=c)
                        cell.border = ExcelEstilos.BORDER_THIN
                        cell.font = ExcelEstilos.FONT_NORMAL
                        # Centrar todo excepto Docente
                        cell.alignment = ExcelEstilos.ALIGN_CENTER if c != 6 else ExcelEstilos.ALIGN_LEFT

                    current_row += 1

                # Merge Vertical para GRUPO (Col 1) y DOCENTE (Col 6)
                # No mergeamos AULA (Col 7) porque cambia por horario
                filas_grupo = current_row - fila_inicio_grupo
                if filas_grupo > 0:
                    fila_fin = current_row - 1

                    # Merge Grupo
                    ws.merge_cells(start_row=fila_inicio_grupo, start_column=1, end_row=fila_fin, end_column=1)

                    # Merge Docente
                    #ws.merge_cells(start_row=fila_inicio_grupo, start_column=6, end_row=fila_fin, end_column=6)
                    #ws.cell(row=fila_inicio_grupo, column=6).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            current_row += 1 # Espacio entre cursos

    # Ajustar Anchos de Columna
    column_widths = {
        'A': 8,  # Grupo
        'B': 5,  # T
        'C': 12, # Dia
        'D': 8,  # Inicio
        'E': 8,  # Fin
        'F': 40, # Docente (Más ancho para que quepa el nombre)
        'G': 12  # Aula
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


# Hoja 2: Horario gráfico
def _generar_hojas_horario_grafico(wb, grupos, facultad_texto, dependencia_texto, nombre_periodo):
    ws = wb.create_sheet(title="Horarios")

    # Aquí el ancho es fijo (7 columnas: Hora + 6 días), así que centramos en 7
    _agregar_cabecera(ws, facultad_texto, dependencia_texto, nombre_periodo, ancho_total=7)

    col_dias = {1: 2, 2: 3, 3: 4, 4: 5, 5: 6, 6: 7}
    dias_label = ["HORA", "LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO"]
    start_hour = 8
    end_hour = 22
    total_horas = end_hour - start_hour

    colores_por_curso = {}

    grupos_por_ciclo = defaultdict(list)
    for g in grupos:
        grupos_por_ciclo[g.asignatura.ciclo].append(g)

    fila_actual = 7

    for ciclo, grupos_ciclo in sorted(grupos_por_ciclo.items()):

        ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=7)
        cell = ws.cell(row=fila_actual, column=1, value=f"CICLO {ciclo}")
        cell.font = ExcelEstilos.FONT_CICLO
        cell.fill = PatternFill(start_color=ExcelEstilos.COLOR_TITULO_CICLO, fill_type="solid")
        cell.alignment = ExcelEstilos.ALIGN_CENTER
        cell.border = ExcelEstilos.BORDER_THIN

        fila_actual += 2

        secciones = defaultdict(list)
        for grp in grupos_ciclo:
            secciones[grp.numero].append(grp)

        for numero_seccion in sorted(secciones.keys()):
            grupos_seccion = secciones[numero_seccion]

            # titulo de grupo
            ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=7)
            cell = ws.cell(row=fila_actual, column=1, value=f"G-{numero_seccion}")
            cell.font = ExcelEstilos.FONT_HEADER
            cell.fill = PatternFill(start_color=ExcelEstilos.COLOR_CABECERA_AZUL, fill_type="solid")
            cell.alignment = ExcelEstilos.ALIGN_CENTER
            cell.border = ExcelEstilos.BORDER_THIN
            fila_actual += 1

            for idx, texto in enumerate(dias_label):
                cell = ws.cell(row=fila_actual, column=idx + 1, value=texto)
                cell.font = ExcelEstilos.FONT_HEADER
                cell.border = ExcelEstilos.BORDER_THIN
                cell.alignment = ExcelEstilos.ALIGN_CENTER

            fila_actual += 1
            fila_inicio_grid = fila_actual

            for h in range(start_hour, end_hour):
                row_idx = fila_inicio_grid + (h - start_hour)
                hora_str = f"{h:02d}:00 - {h+1:02d}:00"

                cell = ws.cell(row=row_idx, column=1, value=hora_str)
                cell.font = ExcelEstilos.FONT_NORMAL
                cell.border = ExcelEstilos.BORDER_THIN
                cell.alignment = ExcelEstilos.ALIGN_CENTER

                for c in range(2, 8):
                    ws.cell(row=row_idx, column=c).border = ExcelEstilos.BORDER_THIN

            bloques_por_columna = defaultdict(list)

            for grp in grupos_seccion:
                for hor in grp.horarios.all():

                    if hor.dia not in col_dias:
                        continue

                    h_ini = hor.hora_inicio.hour
                    h_fin = hor.hora_fin.hour
                    if hor.hora_fin.minute > 0:
                        h_fin += 1

                    if h_ini < start_hour:
                        continue

                    col_idx = col_dias[hor.dia]
                    offset = h_ini - start_hour
                    duracion = h_fin - h_ini

                    row_start = fila_inicio_grid + offset
                    row_end = row_start + duracion - 1

                    curso_nombre = grp.asignatura.nombre

                    if curso_nombre not in colores_por_curso:
                        colores_por_curso[curso_nombre] = generar_color_hex(curso_nombre)

                    color_hex = colores_por_curso[curso_nombre]

                    aula_str = f" - {hor.aula.nombre}" if hor.aula else ""
                    docente = str(hor.docente).split()[0] if hor.docente else ""

                    texto = f"{curso_nombre}\n{docente}{aula_str}"

                    cell = ws.cell(row=row_start, column=col_idx)

                    if cell.value:
                        cell.value = str(cell.value) + "\n\n" + texto
                    else:
                        cell.value = texto

                    cell.font = Font(name='Calibri', size=8, bold=True)
                    cell.alignment = ExcelEstilos.ALIGN_CENTER
                    cell.fill = PatternFill(start_color=color_hex,
                                            end_color=color_hex,
                                            fill_type="solid")


                    bloques_por_columna[col_idx].append({
                        "row_start": row_start,
                        "row_end": row_end,
                        "curso": curso_nombre
                    })

            for col_idx, bloques in bloques_por_columna.items():
                bloques = sorted(bloques, key=lambda x: x["row_start"])

                i = 0
                while i < len(bloques):
                    inicio = bloques[i]["row_start"]
                    fin = bloques[i]["row_end"]
                    curso_actual = bloques[i]["curso"]

                    j = i + 1

                    while (
                        j < len(bloques) and
                        bloques[j]["curso"] == curso_actual and
                        bloques[j]["row_start"] == fin + 1
                    ):
                        fin = bloques[j]["row_end"]
                        j += 1

                    if fin > inicio:
                        try:
                            ws.merge_cells(
                                start_row=inicio,
                                start_column=col_idx,
                                end_row=fin,
                                end_column=col_idx
                            )
                        except:
                            pass

                    i = j

            fila_actual = fila_inicio_grid + total_horas + 2

        fila_actual += 2

    ws.column_dimensions['A'].width = 15
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 25


# Funcion principal para generar el reporte
def generar_reporte_grupos(grupos_queryset, user):
    if not grupos_queryset.exists():
        return None

    # Obtenemos el periodo del primer grupo (ya que todos son del mismo periodo por el filtro)
    first_group = grupos_queryset.first()
    nombre_periodo_str = str(first_group.periodo)

    facultad_texto = user.facultad.nombre.upper() if getattr(user, "facultad", None) else "FACULTAD"
    rol_nombre = getattr(user.rol, "name", "")
    dependencia_texto = ""

    if rol_nombre == "Vicedecano Académico":
        dependencia_texto = "VICEDECANATO ACADÉMICO"
        escuela_nombre = first_group.asignatura.plan.escuela.nombre.upper()
        dependencia_texto += f" - ESCUELA PROFESIONAL DE {escuela_nombre}"
    elif rol_nombre in ["Coordinador de Estudios Generales", "Jefe de Estudios Generales"]:
        dependencia_texto = "COORDINACIÓN DE ESTUDIOS GENERALES"
    elif getattr(user, "escuela", None):
        dependencia_texto = f"ESCUELA PROFESIONAL DE {user.escuela.nombre.upper()}"

    grupos = list(grupos_queryset)

    wb = Workbook()

    # Pasamos los datos a tus funciones de pintado (sin cambios internos en ellas)
    _generar_hoja_lista_cursos(wb, grupos, facultad_texto, dependencia_texto, nombre_periodo_str)
    _generar_hojas_horario_grafico(wb, grupos, facultad_texto, dependencia_texto, nombre_periodo_str)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    periodo_slug = slugify(nombre_periodo_str)
    dependencia_slug = slugify(dependencia_texto)

    filename = f"horarios-{periodo_slug}-{dependencia_slug}.xlsx"

    return output, filename
