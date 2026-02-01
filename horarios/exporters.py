from io import BytesIO
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from asignaturas.models import Asignatura
from grupos.models import Grupo
from periodos.models import PeriodoAcademico
from .excel_styles import ExcelEstilos, crear_rango_horas, aplicar_estilo_celda, generar_color_asignatura


def generar_reporte_grupos(periodo_id=None, grupos_ids=None, archivo_salida=None):
    # Obtener periodo
    if periodo_id:
        periodo = PeriodoAcademico.objects.get(id=periodo_id)
    else:
        periodo = PeriodoAcademico.objects.filter(
            fecha_inicio__lte=datetime.now().date(),
            fecha_fin__gte=datetime.now().date()
        ).first()
    
    # Obtener grupos
    if grupos_ids:
        grupos = Grupo.objects.filter(id__in=grupos_ids).select_related('asignatura', 'periodo', 'docente').prefetch_related('horarios', 'horarios__aula').order_by('nombre')
    elif periodo:
        grupos = Grupo.objects.filter(periodo=periodo).select_related('asignatura', 'periodo', 'docente').prefetch_related('horarios', 'horarios__aula').order_by('nombre')
    else:
        grupos = Grupo.objects.all().select_related('asignatura', 'periodo', 'docente').prefetch_related('horarios', 'horarios__aula').order_by('nombre')[:10]
    
    # Agrupar grupos por ciclo de la asignatura
    grupos_por_ciclo = {}
    for grupo in grupos:
        ciclo = grupo.asignatura.ciclo if hasattr(grupo.asignatura, 'ciclo') else 1
        if ciclo not in grupos_por_ciclo:
            grupos_por_ciclo[ciclo] = {}
        
        nombre_grupo = grupo.nombre
        if nombre_grupo not in grupos_por_ciclo[ciclo]:
            grupos_por_ciclo[ciclo][nombre_grupo] = []
        grupos_por_ciclo[ciclo][nombre_grupo].append(grupo)
    
    # Crear workbook
    wb = Workbook()
    primera_hoja = True
    
    # Crear una hoja por ciclo
    for ciclo in sorted(grupos_por_ciclo.keys()):
        grupos_dict = grupos_por_ciclo[ciclo]
        
        if primera_hoja:
            ws = wb.active
            ws.title = f"CICLO {ciclo}"
            primera_hoja = False
        else:
            ws = wb.create_sheet(f"CICLO {ciclo}")
        
        fila_actual = 1
        
        # Por cada grupo, crear su sección
        for nombre_grupo, grupos_del_grupo in sorted(grupos_dict.items()):
            # Título del grupo (fila amarilla)
            ws.merge_cells(f'A{fila_actual}:N{fila_actual}')
            titulo_celda = ws[f'A{fila_actual}']
            aplicar_estilo_celda(
                titulo_celda,
                f"GRUPO {nombre_grupo.replace('G-', '')}",
                font=Font(name='Arial', size=12, bold=True),
                fill=PatternFill(start_color=ExcelEstilos.COLOR_CABECERA, fill_type='solid'),
                alignment=ExcelEstilos.ALIGN_CENTER,
                border=ExcelEstilos.BORDER_THIN
            )
            fila_actual += 1
            
            # Cabecera de días (con columna T entre cada día)
            # Estructura: H. INI. | H. FIN | T | LUNES | T | MARTES | T | MIERCOLES | T | JUEVES | T | VIERNES | T | SABADO
            cabeceras = ['H. INI.', 'H. FIN', 'T', 'LUNES', 'T', 'MARTES', 'T', 'MIERCOLES', 'T', 'JUEVES', 'T', 'VIERNES', 'T', 'SABADO']
            
            for col_idx, cabecera in enumerate(cabeceras, 1):
                celda = ws.cell(row=fila_actual, column=col_idx)
                aplicar_estilo_celda(
                    celda,
                    cabecera,
                    font=ExcelEstilos.FONT_CABECERA,
                    fill=PatternFill(start_color=ExcelEstilos.COLOR_CABECERA, fill_type='solid'),
                    alignment=ExcelEstilos.ALIGN_CENTER,
                    border=ExcelEstilos.BORDER_THIN
                )
            fila_actual += 1
            
            # Crear rangos de horario (08:00 a 22:00)
            rangos_horas = crear_rango_horas(hora_inicio=time(8, 0), hora_fin=time(22, 0), intervalo_minutos=60)
            
            # Crear matriz de horarios
            # Estructura: horarios_matriz[fila][dia] = [(asignatura, tipo, color), ...]
            horarios_matriz = {}
            
            # Recopilar TODOS los horarios de TODOS los grupos con este nombre
            for grupo in grupos_del_grupo:
                for horario in grupo.horarios.all():
                    # Encontrar la fila correspondiente
                    for idx, rango in enumerate(rangos_horas):
                        if horario.hora_inicio == rango['inicio']:
                            fila_idx = idx
                            dia_col = horario.dia  # 1=Lunes, 2=Martes, etc.
                            
                            # Calcular duración en horas
                            duracion = (datetime.combine(datetime.today(), horario.hora_fin) - 
                                      datetime.combine(datetime.today(), horario.hora_inicio)).seconds / 3600
                            num_filas = int(duracion)
                            
                            # Generar color único para esta asignatura
                            color_asignatura = generar_color_asignatura(grupo.asignatura.nombre)
                            
                            # Guardar en la matriz
                            for f in range(fila_idx, fila_idx + num_filas):
                                if f not in horarios_matriz:
                                    horarios_matriz[f] = {}
                                if dia_col not in horarios_matriz[f]:
                                    horarios_matriz[f][dia_col] = []
                                
                                horarios_matriz[f][dia_col].append({
                                    'asignatura': grupo.asignatura.nombre,
                                    'tipo': horario.tipo,
                                    'color': color_asignatura,
                                    'inicio': fila_idx == f,  # True solo en la primera fila
                                    'filas': num_filas
                                })
                            break
            
            # Llenar la grilla
            for idx, rango in enumerate(rangos_horas):
                fila = fila_actual + idx
                
                # Columnas de hora
                ws.cell(row=fila, column=1).value = rango['inicio'].strftime('%H:%M')
                ws.cell(row=fila, column=2).value = rango['fin'].strftime('%H:%M')
                
                for col in [1, 2]:
                    celda = ws.cell(row=fila, column=col)
                    celda.fill = PatternFill(start_color=ExcelEstilos.COLOR_HORA, fill_type='solid')
                    celda.alignment = ExcelEstilos.ALIGN_CENTER
                    celda.border = ExcelEstilos.BORDER_THIN
                    celda.font = Font(name='Arial', size=9)
                
                for dia in range(1, 7):  # 1=Lunes a 6=Sábado
                    col_t = 3 + (dia - 1) * 2  # Columna T
                    col_asig = col_t + 1  # Columna asignatura
                    
                    if idx in horarios_matriz and dia in horarios_matriz[idx]:
                        # Hay horario(s) en este día/hora
                        horarios_dia = horarios_matriz[idx][dia]
                        
                        # Tomar el primer horario (si hay múltiples, se sobrepondrán)
                        for info in horarios_dia:
                            if info['inicio']:  # Solo escribir en la primera fila del merge
                                # Usar el color de la asignatura
                                color_celda = info['color']
                                
                                # Columna T
                                celda_tipo = ws.cell(row=fila, column=col_t)
                                aplicar_estilo_celda(
                                    celda_tipo,
                                    info['tipo'],
                                    font=Font(name='Arial', size=10, bold=True),
                                    fill=PatternFill(start_color=color_celda, fill_type='solid'),
                                    alignment=ExcelEstilos.ALIGN_CENTER,
                                    border=ExcelEstilos.BORDER_THIN
                                )
                                
                                # Celda de asignatura
                                celda_asig = ws.cell(row=fila, column=col_asig)
                                aplicar_estilo_celda(
                                    celda_asig,
                                    info['asignatura'],
                                    font=Font(name='Arial', size=9, bold=False),
                                    fill=PatternFill(start_color=color_celda, fill_type='solid'),
                                    alignment=ExcelEstilos.ALIGN_CENTER,
                                    border=ExcelEstilos.BORDER_THIN
                                )
                                
                                # Merge si ocupa más de una fila
                                if info['filas'] > 1:
                                    try:
                                        ws.merge_cells(
                                            start_row=fila,
                                            start_column=col_t,
                                            end_row=fila + info['filas'] - 1,
                                            end_column=col_t
                                        )
                                        ws.merge_cells(
                                            start_row=fila,
                                            start_column=col_asig,
                                            end_row=fila + info['filas'] - 1,
                                            end_column=col_asig
                                        )
                                    except:
                                        pass  # Si ya está merged, continuar
                    else:
                        # Celda vacía
                        celda_t = ws.cell(row=fila, column=col_t)
                        celda_t.border = ExcelEstilos.BORDER_THIN
                        
                        celda_asig = ws.cell(row=fila, column=col_asig)
                        celda_asig.border = ExcelEstilos.BORDER_THIN
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 7
            
            # Columnas T (estrechas)
            for col_letra in ['C', 'E', 'G', 'I', 'K', 'M']:
                ws.column_dimensions[col_letra].width = 3
            
            # Columnas de días (anchas)
            for col_letra in ['D', 'F', 'H', 'J', 'L', 'N']:
                ws.column_dimensions[col_letra].width = 20
            
            # Espacio entre grupos
            fila_actual += len(rangos_horas) + 2
    
    # Guardar
    if archivo_salida:
        wb.save(archivo_salida)
        return None
    else:
        archivo_bytes = BytesIO()
        wb.save(archivo_bytes)
        archivo_bytes.seek(0)
        return archivo_bytes


def generar_reporte_por_asignatura(periodo_id=None, ciclo=None, archivo_salida=None):
    """
    Genera un reporte de grupos y horarios agrupados por asignatura y ciclo.
    
    Args:
        periodo_id: ID del periodo académico
        ciclo: Ciclo específico (si es None, incluye todos)
        archivo_salida: Ruta del archivo de salida (si es None, retorna BytesIO)
    
    Returns:
        BytesIO o None (si se especifica archivo_salida)
    """
    # Obtener periodo
    if periodo_id:
        periodo = PeriodoAcademico.objects.get(id=periodo_id)
    else:
        periodo = PeriodoAcademico.objects.filter(
            fecha_inicio__lte=datetime.now().date(),
            fecha_fin__gte=datetime.now().date()
        ).first()
    
    # Obtener asignaturas
    asignaturas_query = Asignatura.objects.all()
    if ciclo:
        asignaturas_query = asignaturas_query.filter(ciclo=ciclo)
    
    asignaturas = asignaturas_query.order_by('ciclo', 'nombre')
    
    # Crear workbook
    wb = Workbook()
    
    # Agrupar por ciclo
    ciclos = {}
    for asignatura in asignaturas:
        if asignatura.ciclo not in ciclos:
            ciclos[asignatura.ciclo] = []
        ciclos[asignatura.ciclo].append(asignatura)
    
    # Crear una hoja por ciclo
    primera_hoja = True
    for ciclo_num, asignaturas_ciclo in sorted(ciclos.items()):
        if primera_hoja:
            ws = wb.active
            ws.title = f"CICLO {ciclo_num}"
            primera_hoja = False
        else:
            ws = wb.create_sheet(f"CICLO {ciclo_num}")
        
        # Título del ciclo
        ws.merge_cells('A1:F1')
        titulo_celda = ws['A1']
        aplicar_estilo_celda(
            titulo_celda,
            f"CICLO {ciclo_num}",
            font=ExcelEstilos.FONT_TITULO,
            fill=PatternFill(start_color=ExcelEstilos.COLOR_CABECERA, fill_type='solid'),
            alignment=ExcelEstilos.ALIGN_CENTER,
            border=ExcelEstilos.BORDER_THIN
        )
        
        fila_actual = 3
        
        # Por cada asignatura del ciclo
        for asignatura in asignaturas_ciclo:
            # Título de asignatura
            ws.merge_cells(f'A{fila_actual}:F{fila_actual}')
            celda_asignatura = ws[f'A{fila_actual}']
            aplicar_estilo_celda(
                celda_asignatura,
                asignatura.nombre.upper(),
                font=Font(name='Arial', size=12, bold=True),
                fill=PatternFill(start_color="D9D9D9", fill_type='solid'),
                alignment=ExcelEstilos.ALIGN_CENTER,
                border=ExcelEstilos.BORDER_THIN
            )
            fila_actual += 1
            
            # Cabecera de tabla - HORARIO dividido en 2 columnas
            cabeceras = ['GRUPO', 'T', 'DIA', 'HORARIO', 'DOCENTE PROPUESTO']
            col_actual = 1
            
            for cabecera in cabeceras:
                celda = ws.cell(row=fila_actual, column=col_actual)
                
                if cabecera == 'HORARIO':
                    # Merge HORARIO sobre 2 columnas
                    ws.merge_cells(
                        start_row=fila_actual,
                        start_column=col_actual,
                        end_row=fila_actual,
                        end_column=col_actual + 1
                    )
                    aplicar_estilo_celda(
                        celda,
                        cabecera,
                        font=ExcelEstilos.FONT_CABECERA,
                        fill=PatternFill(start_color=ExcelEstilos.COLOR_CABECERA, fill_type='solid'),
                        alignment=ExcelEstilos.ALIGN_CENTER,
                        border=ExcelEstilos.BORDER_THIN
                    )
                    # También aplicar estilo a la segunda celda del merge
                    celda_2 = ws.cell(row=fila_actual, column=col_actual + 1)
                    celda_2.fill = PatternFill(start_color=ExcelEstilos.COLOR_CABECERA, fill_type='solid')
                    celda_2.border = ExcelEstilos.BORDER_THIN
                    col_actual += 2  # Saltar 2 columnas
                else:
                    aplicar_estilo_celda(
                        celda,
                        cabecera,
                        font=ExcelEstilos.FONT_CABECERA,
                        fill=PatternFill(start_color=ExcelEstilos.COLOR_CABECERA, fill_type='solid'),
                        alignment=ExcelEstilos.ALIGN_CENTER,
                        border=ExcelEstilos.BORDER_THIN
                    )
                    col_actual += 1
            fila_actual += 1
            
            # Obtener grupos de esta asignatura
            grupos_query = Grupo.objects.filter(
                asignatura=asignatura
            ).select_related('docente').prefetch_related('horarios')
            
            # Filtrar por periodo solo si existe
            if periodo:
                grupos_query = grupos_query.filter(periodo=periodo)
            
            grupos = grupos_query
            
            if not grupos.exists():
                # Si no hay grupos, dejar una fila vacía
                ws.cell(row=fila_actual, column=1).value = "Sin grupos"
                fila_actual += 1
            else:
                # Por cada grupo
                for grupo in grupos:
                    horarios = grupo.horarios.all().order_by('tipo', 'dia', 'hora_inicio')
                    
                    if not horarios.exists():
                        # Grupo sin horarios
                        ws.cell(row=fila_actual, column=1).value = grupo.nombre
                        ws.cell(row=fila_actual, column=5).value = str(grupo.docente) if grupo.docente else ""
                        fila_actual += 1
                    else:
                        fila_inicio_grupo = fila_actual
                        
                        # Por cada horario del grupo
                        for horario in horarios:
                            ws.cell(row=fila_actual, column=1).value = grupo.nombre
                            ws.cell(row=fila_actual, column=2).value = horario.tipo
                            ws.cell(row=fila_actual, column=3).value = horario.get_dia_display()
                            # HORARIO dividido en 2 columnas: inicio y fin
                            ws.cell(row=fila_actual, column=4).value = horario.hora_inicio.strftime('%H:%M')
                            ws.cell(row=fila_actual, column=5).value = horario.hora_fin.strftime('%H:%M')
                            ws.cell(row=fila_actual, column=6).value = str(grupo.docente) if grupo.docente else ""
                            
                            # Aplicar estilos
                            for col in range(1, 7):
                                celda = ws.cell(row=fila_actual, column=col)
                                celda.border = ExcelEstilos.BORDER_THIN
                                celda.font = ExcelEstilos.FONT_NORMAL
                                celda.alignment = ExcelEstilos.ALIGN_CENTER if col <= 5 else ExcelEstilos.ALIGN_LEFT
                            
                            fila_actual += 1
                        
                        # Merge celdas de grupo y docente si hay múltiples horarios
                        if len(horarios) > 1:
                            ws.merge_cells(f'A{fila_inicio_grupo}:A{fila_actual-1}')
                            ws.merge_cells(f'F{fila_inicio_grupo}:F{fila_actual-1}')
            
            fila_actual += 1  # Espacio entre asignaturas
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10  # Hora inicio
        ws.column_dimensions['E'].width = 10  # Hora fin
        ws.column_dimensions['F'].width = 30  # Docente
    
    # Guardar
    if archivo_salida:
        wb.save(archivo_salida)
        return None
    else:
        archivo_bytes = BytesIO()
        wb.save(archivo_bytes)
        archivo_bytes.seek(0)
        return archivo_bytes


def generar_reporte_matricula(periodo_id=None, archivo_salida=None):
    """
    Genera un reporte completo de matrícula con información de grupos,
    vacantes, matriculados y horarios.
    
    Args:
        periodo_id: ID del periodo académico
        archivo_salida: Ruta del archivo de salida (si es None, retorna BytesIO)
    
    Returns:
        BytesIO o None (si se especifica archivo_salida)
    """
    # Obtener periodo
    if periodo_id:
        periodo = PeriodoAcademico.objects.get(id=periodo_id)
    else:
        periodo = PeriodoAcademico.objects.filter(
            fecha_inicio__lte=datetime.now().date(),
            fecha_fin__gte=datetime.now().date()
        ).first()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Matrícula"
    
    # Título
    ws.merge_cells('A1:H1')
    titulo_celda = ws['A1']
    aplicar_estilo_celda(
        titulo_celda,
        f"REPORTE DE MATRÍCULA - {periodo.nombre if periodo else 'TODOS'}",
        font=ExcelEstilos.FONT_TITULO,
        fill=PatternFill(start_color=ExcelEstilos.COLOR_CABECERA, fill_type='solid'),
        alignment=ExcelEstilos.ALIGN_CENTER,
        border=ExcelEstilos.BORDER_THIN
    )
    
    # Cabeceras
    cabeceras = ['ASIGNATURA', 'GRUPO', 'DOCENTE', 'VACANTES', 'MATRICULADOS', 'DISPONIBLES', 'DIA', 'HORARIO']
    fila_cabecera = 3
    
    for col_idx, cabecera in enumerate(cabeceras, 1):
        celda = ws.cell(row=fila_cabecera, column=col_idx)
        aplicar_estilo_celda(
            celda,
            cabecera,
            font=ExcelEstilos.FONT_CABECERA,
            fill=PatternFill(start_color=ExcelEstilos.COLOR_CABECERA, fill_type='solid'),
            alignment=ExcelEstilos.ALIGN_CENTER,
            border=ExcelEstilos.BORDER_THIN
        )
    
    # Obtener grupos
    grupos_query = Grupo.objects.select_related(
        'asignatura', 'docente', 'periodo'
    ).prefetch_related('horarios', 'vacantes')
    
    if periodo:
        grupos_query = grupos_query.filter(periodo=periodo)
    
    grupos = grupos_query.order_by('asignatura__ciclo', 'asignatura__nombre', 'nombre')
    
    fila_actual = fila_cabecera + 1
    
    for grupo in grupos:
        vacantes_total = grupo.total_vacantes
        # Aquí podrías calcular matriculados desde la BD si tuvieras ese campo
        matriculados = sum(v.matriculados for v in grupo.vacantes.all())
        disponibles = vacantes_total - matriculados
        
        horarios = grupo.horarios.all().order_by('dia', 'hora_inicio')
        
        if not horarios.exists():
            ws.cell(row=fila_actual, column=1).value = grupo.asignatura.nombre
            ws.cell(row=fila_actual, column=2).value = grupo.nombre
            ws.cell(row=fila_actual, column=3).value = str(grupo.docente) if grupo.docente else ""
            ws.cell(row=fila_actual, column=4).value = vacantes_total
            ws.cell(row=fila_actual, column=5).value = matriculados
            ws.cell(row=fila_actual, column=6).value = disponibles
            fila_actual += 1
        else:
            fila_inicio = fila_actual
            
            for horario in horarios:
                ws.cell(row=fila_actual, column=1).value = grupo.asignatura.nombre
                ws.cell(row=fila_actual, column=2).value = grupo.nombre
                ws.cell(row=fila_actual, column=3).value = str(grupo.docente) if grupo.docente else ""
                ws.cell(row=fila_actual, column=4).value = vacantes_total
                ws.cell(row=fila_actual, column=5).value = matriculados
                ws.cell(row=fila_actual, column=6).value = disponibles
                ws.cell(row=fila_actual, column=7).value = horario.get_dia_display()
                ws.cell(row=fila_actual, column=8).value = f"{horario.hora_inicio.strftime('%H:%M')} - {horario.hora_fin.strftime('%H:%M')}"
                
                # Aplicar estilos
                for col in range(1, 9):
                    celda = ws.cell(row=fila_actual, column=col)
                    celda.border = ExcelEstilos.BORDER_THIN
                    celda.font = ExcelEstilos.FONT_NORMAL
                    celda.alignment = ExcelEstilos.ALIGN_CENTER
                
                fila_actual += 1
            
            # Merge celdas de información del grupo
            if len(horarios) > 1:
                for col in range(1, 7):
                    ws.merge_cells(
                        start_row=fila_inicio,
                        start_column=col,
                        end_row=fila_actual - 1,
                        end_column=col
                    )
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    
    # Guardar
    if archivo_salida:
        wb.save(archivo_salida)
        return None
    else:
        archivo_bytes = BytesIO()
        wb.save(archivo_bytes)
        archivo_bytes.seek(0)
        return archivo_bytes
