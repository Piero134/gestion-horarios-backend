from datetime import datetime, date
import re
from openpyxl import load_workbook
from django.core.exceptions import ValidationError
from django.db import transaction

# --- IMPORTACIONES DE MODELOS ---
from asignaturas.models import Asignatura
from planes.models import PlanEstudios
from escuelas.models import Escuela
from facultades.models import Facultad
from grupos.models import Grupo, DistribucionVacantes
from periodos.models import PeriodoAcademico
from .models import Horario
from .utils import (
    limpiar_texto,
    interpretar_dia,
    interpretar_hora,
    interpretar_tipo_sesion,
    obtener_o_crear_aula,
    obtener_o_crear_docente
)

def importar_horarios_desde_excel(archivo_excel, periodo_id=None):
    workbook = load_workbook(archivo_excel)
    
    # ==========================================
    # 1. PREPARACIÓN DE DATOS PADRE
    # ==========================================
    
    print("   -> Verificando estructura base (Facultad/Escuela/Plan)...")
    
    # 1. Crear Facultad (Código numérico para evitar ValueError)
    facultad, _ = Facultad.objects.get_or_create(
        nombre="FACULTAD DE INGENIERÍA Y SISTEMAS",
        defaults={'codigo': 200}
    )

    # 2. Crear Escuela (Código numérico)
    escuela, _ = Escuela.objects.get_or_create(
        nombre="INGENIERÍA DE SOFTWARE",
        defaults={'facultad': facultad, 'codigo': 201}
    )

    # 3. Crear Plan de Estudios (Solo nombre y escuela, sin codigo ni anio)
    plan_default, _ = PlanEstudios.objects.get_or_create(
        nombre="Plan Importado 2026", 
        escuela=escuela
    )

    # ==========================================
    # 2. GESTIÓN DEL PERIODO
    # ==========================================
    if periodo_id:
        periodo = PeriodoAcademico.objects.get(id=periodo_id)
    else:
        anio_actual = datetime.now().year
        # CORREGIDO: Eliminamos 'activo' del defaults
        periodo, _ = PeriodoAcademico.objects.get_or_create(
            nombre=f"{anio_actual}-AUTO",
            defaults={
                'tipo': 'SEMESTRE',
                'anio': anio_actual,
                'fecha_inicio': date(anio_actual, 3, 1),
                'fecha_fin': date(anio_actual, 7, 31)
            }
        )

    grupos_creados = 0
    horarios_creados = 0
    errores = []
    
    # ==========================================
    # 3. PROCESAMIENTO DEL EXCEL
    # ==========================================
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"   -> Procesando hoja: {sheet_name}")
        try:
            # Buscar encabezados
            header_row = None
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15), 1):
                cell_values = [limpiar_texto(cell.value).upper() for cell in row]
                if 'GRUPO' in cell_values or 'ASIGNATURA' in cell_values:
                    header_row = row_idx
                    break
            
            if not header_row:
                errores.append(f"Hoja '{sheet_name}': No se encontró fila de encabezados.")
                continue
            
            # Mapeo de columnas
            headers = {}
            for col_idx, cell in enumerate(sheet[header_row], 1):
                txt = limpiar_texto(cell.value).upper()
                if 'ASIGNATURA' in txt: headers['asignatura'] = col_idx
                elif txt in ['GL', 'GRUPO', 'G']: headers['grupo'] = col_idx
                elif txt in ['T/P/L', 'T', 'TIPO']: headers['tipo'] = col_idx
                elif 'DIA' in txt or 'DÍA' in txt: headers['dia'] = col_idx
                elif 'H.INI' in txt or 'INICIO' in txt: headers['hora_inicio'] = col_idx
                elif 'H.FIN' in txt or 'FIN' in txt: headers['hora_fin'] = col_idx
                elif 'DOCENTE' in txt: headers['docente'] = col_idx
                elif 'AULA' in txt: headers['aula'] = col_idx
            
            # Procesar filas
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                row = sheet[row_idx]
                
                nombre_asignatura = limpiar_texto(row[headers.get('asignatura', 1) - 1].value) if 'asignatura' in headers else None
                nombre_grupo = limpiar_texto(row[headers.get('grupo', 2) - 1].value) if 'grupo' in headers else None
                
                if not nombre_asignatura or not nombre_grupo:
                    continue

                try:
                    # --- A. ASIGNATURA ---
                    asignatura = Asignatura.objects.filter(nombre__iexact=nombre_asignatura, plan=plan_default).first()
                    
                    if not asignatura:
                        # Crear asignatura con datos suficientes para validaciones (horas > 0)
                        codigo_gen = nombre_asignatura[:3].upper() + f"-{row_idx}"
                        
                        asignatura = Asignatura.objects.create(
                            plan=plan_default,
                            nombre=nombre_asignatura,
                            codigo=codigo_gen,
                            ciclo=1,            
                            creditos=4,         
                            tipo='O',           
                            horas_teoria=6,      
                            horas_practica=6,
                            horas_laboratorio=6
                        )

                    # --- B. GRUPO ---
                    nombre_docente = limpiar_texto(row[headers.get('docente', 8) - 1].value) if 'docente' in headers else None
                    docente = obtener_o_crear_docente(nombre_docente) if nombre_docente else None
                    
                    grupo, g_created = Grupo.objects.get_or_create(
                        nombre=nombre_grupo,
                        asignatura=asignatura,
                        periodo=periodo,
                        defaults={'docente': docente}
                    )
                    
                    if g_created:
                        grupos_creados += 1
                        # Crear vacantes (usamos la misma asignatura para evitar error de Equivalencia)
                        DistribucionVacantes.objects.create(
                            grupo=grupo,
                            asignatura=asignatura, 
                            cantidad=30
                        )

                    # --- C. HORARIO ---
                    dia_txt = limpiar_texto(row[headers.get('dia', 4) - 1].value) if 'dia' in headers else None
                    h_ini_val = row[headers['hora_inicio'] - 1].value if 'hora_inicio' in headers else None
                    h_fin_val = row[headers['hora_fin'] - 1].value if 'hora_fin' in headers else None
                    
                    if dia_txt and h_ini_val and h_fin_val:
                        dia_num = interpretar_dia(dia_txt)
                        hora_inicio = interpretar_hora(h_ini_val)
                        hora_fin = interpretar_hora(h_fin_val)
                        
                        if dia_num and hora_inicio and hora_fin:
                            tipo_sesion_txt = limpiar_texto(row[headers.get('tipo', 3) - 1].value) if 'tipo' in headers else 'T'
                            tipo_sesion = interpretar_tipo_sesion(tipo_sesion_txt)
                            
                            aula_nombre = limpiar_texto(row[headers.get('aula', 9) - 1].value) if 'aula' in headers else f"AULA-{nombre_grupo}"
                            aula = obtener_o_crear_aula(aula_nombre, es_laboratorio=(tipo_sesion == 'L'))

                            Horario.objects.get_or_create(
                                grupo=grupo,
                                dia=dia_num,
                                hora_inicio=hora_inicio,
                                hora_fin=hora_fin,
                                defaults={
                                    'tipo': tipo_sesion,
                                    'aula': aula
                                }
                            )
                            horarios_creados += 1

                except Exception as e:
                    errores.append(f"Fila {row_idx} ({nombre_asignatura}): {str(e)}")

        except Exception as e:
            errores.append(f"Error fatal en hoja '{sheet_name}': {str(e)}")
    
    return {
        'success': True,
        'grupos_creados': grupos_creados,
        'horarios_creados': horarios_creados,
        'errores': errores,
        'periodo': str(periodo)
    }