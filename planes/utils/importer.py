import openpyxl
import unicodedata
from django.db import transaction
from django.core.exceptions import ValidationError
from asignaturas.models import Asignatura, Prerequisito

def normalizar_texto(texto):
    if not texto: return ""
    texto = str(texto).strip().upper()
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )

def safe_int(valor):
    if valor is None or str(valor).strip() == "": return 0
    try:
        return int(float(valor))
    except (ValueError, TypeError):
        return 0

def procesar_excel_plan(plan_obj, excel_file):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb.active

    columnas_obligatorias = {'CICLO', 'TIPO', 'CODIGO', 'NOMBRE', 'CREDITOS', 'PREREQUISITO'}
    columnas_opcionales = {'HT', 'HP', 'HL'}

    header_map = {} # Guarda { 'NOMBRE_COLUMNA': INDICE }
    header_row_index = None

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        # Limpiamos y normalizamos los valores de la fila actual
        fila_normalizada = [normalizar_texto(cell) for cell in row if cell is not None]

        if columnas_obligatorias.issubset(set(fila_normalizada)):
            header_row_index = row_idx
            # Mapeamos cada columna encontrada a su índice (0-based)
            for cell_idx, cell_value in enumerate(row):
                val_norm = normalizar_texto(cell_value)
                if val_norm:
                    header_map[val_norm] = cell_idx
            break

    if not header_row_index:
        columnas_encontradas = set(header_map.keys())
        faltantes = columnas_obligatorias - columnas_encontradas
        mensaje_error = (
            f"Faltan las siguientes columnas obligatorias: {', '.join(faltantes)}. "
        )
        raise ValueError(mensaje_error)

    lista_relaciones = []
    asignaturas_creadas = 0

    with transaction.atomic():
        for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            def get_val(key):
                idx = header_map.get(key)
                return row[idx] if idx is not None else None

            codigo = str(get_val('CODIGO') or '').strip()
            # Si la fila está vacía o no tiene código, saltamos
            if not codigo or codigo == 'None' or codigo == "":
                continue

            nombre = normalizar_texto(get_val('NOMBRE') or 'SIN NOMBRE')

            # Crear o actualizar Asignatura
            Asignatura.objects.update_or_create(
                codigo=codigo,
                plan=plan_obj,
                defaults={
                    'nombre': nombre,
                    'ciclo': safe_int(get_val('CICLO')),
                    'creditos': safe_int(get_val('CREDITOS')),
                    'tipo': str(get_val('TIPO') or 'O').strip()[:1].upper(),
                    'horas_teoria': safe_int(get_val('HT')),
                    'horas_practica': safe_int(get_val('HP')),
                    'horas_laboratorio': safe_int(get_val('HL')),
                }
            )
            asignaturas_creadas += 1

            # Guardar prerequisito para la segunda pasada
            req_raw = get_val('PREREQUISITO')
            if req_raw and str(req_raw).strip():
                req_code = str(req_raw).strip().split('-')[0].strip()
                lista_relaciones.append((codigo, req_code))

        mapa_asig = {a.codigo: a for a in Asignatura.objects.filter(plan=plan_obj)}
        Prerequisito.objects.filter(asignatura__plan=plan_obj).delete()

        rels_creadas = 0
        for cod_asig, cod_req in lista_relaciones:
            asig_obj = mapa_asig.get(cod_asig)
            req_obj = mapa_asig.get(cod_req)

            if asig_obj and req_obj:
                try:
                    Prerequisito.objects.create(asignatura=asig_obj, prerequisito=req_obj)
                    rels_creadas += 1
                except ValidationError:
                    pass

    return asignaturas_creadas, rels_creadas
